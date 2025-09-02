"""
Enhanced Excel Processor with Table-Aware Processing
Processes Excel documents sheet by sheet while keeping entire tables together
"""

import openpyxl
from io import BytesIO
import logging
from typing import Dict, Any, List, Optional, Tuple
import datetime
import re
from dataclasses import dataclass

@dataclass
class TableRegion:
    """Represents a table region in an Excel sheet"""
    start_row: int
    end_row: int
    start_col: int
    end_col: int
    headers: List[str]
    data_rows: List[List[str]]
    table_type: str  # 'data_table', 'summary_table', 'list', 'single_values'

class SheetClientDetector:
    """Detect client names in Excel sheet titles"""
    
    def __init__(self):
        # Pattern for client names in sheet titles (matching Autobahn pattern)
        self.client_sheet_patterns = [
            # Direct client name patterns
            re.compile(r'^([A-Za-z\s&\-\.]+?)\s*(?:\s*-\s*|\s+)(exec|executive|contacts?|team|leadership|board).*?$', re.IGNORECASE),
            re.compile(r'^([A-Za-z\s&\-\.]+?)\s*(?:\s*-\s*|\s+)(meeting|agenda|notes|prep).*?$', re.IGNORECASE),
            re.compile(r'^([A-Za-z\s&\-\.]+?)\s*(?:\s*-\s*|\s+)(info|information|details).*?$', re.IGNORECASE),
            # Client name with PM notation
            re.compile(r'^([A-Za-z\s&\-\.]+?)\s*\(PM-([A-Z])\).*?$', re.IGNORECASE),
            # Sheet names that start with likely company names
            re.compile(r'^([A-Z][A-Za-z\s&\-\.]{3,}(?:\s(?:Inc|LLC|Corp|Company|Co|Ltd|Group|Solutions|Services|Systems|Technologies|Tech))).*?$')
        ]
        
        # Common non-client sheet names to exclude
        self.excluded_sheet_names = {
            'summary', 'overview', 'dashboard', 'template', 'instructions', 
            'legend', 'notes', 'data', 'sheet1', 'sheet2', 'sheet3',
            'main', 'index', 'contents', 'home', 'menu', 'config',
            'settings', 'help', 'about', 'version', 'changelog',
            'new template'  # Add this common template name
        }
        
        # PM name mappings (consistent with main system)
        self.pm_names = {
            'C': 'Caleb',
            'S': 'Sam', 
            'K': 'Katherine'
        }
    
    def detect_client_from_sheet_name(self, sheet_name: str) -> Optional[Dict[str, Any]]:
        """Detect client information from sheet name"""
        if not sheet_name or sheet_name.lower().strip() in self.excluded_sheet_names:
            return None
            
        sheet_name = sheet_name.strip()
        
        # Try each pattern
        for pattern in self.client_sheet_patterns:
            match = pattern.match(sheet_name)
            if match:
                potential_client = match.group(1).strip()
                
                # Clean up the client name
                potential_client = self._clean_client_name(potential_client)
                
                # Validate it looks like a real client name
                if self._is_valid_client_name(potential_client):
                    result = {
                        "sheet_client_name": potential_client,
                        "sheet_client_source": "sheet_title",
                        "original_sheet_name": sheet_name,
                        "confidence": self._calculate_confidence(sheet_name, potential_client)
                    }
                    
                    # Check for PM notation
                    if len(match.groups()) > 1 and match.group(2):
                        pm_initial = match.group(2).upper()
                        result.update({
                            "sheet_pm_initial": pm_initial,
                            "sheet_pm_name": self.pm_names.get(pm_initial, "Unknown")
                        })
                    
                    return result
        
        # Fallback: if sheet name looks like a simple company name, use it directly
        if self._is_valid_client_name(sheet_name) and len(sheet_name) >= 3:
            return {
                "sheet_client_name": sheet_name,
                "sheet_client_source": "sheet_title_simple",
                "original_sheet_name": sheet_name,
                "confidence": 0.8  # High confidence for simple names
            }
        
        return None
    
    def _clean_client_name(self, name: str) -> str:
        """Clean and normalize client name"""
        # Remove common prefixes/suffixes
        name = re.sub(r'\b(client|customer|company|corp|corporation|inc|llc|ltd|group)\b', '', name, flags=re.IGNORECASE)
        # Clean up whitespace and punctuation
        name = re.sub(r'\s+', ' ', name).strip()
        name = name.strip(' -.,')
        return name
    
    def _is_valid_client_name(self, name: str) -> bool:
        """Validate if this looks like a real client name"""
        if not name or len(name) < 2:
            return False
        
        # Must contain at least one letter
        if not re.search(r'[A-Za-z]', name):
            return False
            
        # Shouldn't be too short or too long
        if len(name) < 3 or len(name) > 50:
            return False
            
        # Shouldn't be all numbers or common generic terms
        generic_terms = {'test', 'sample', 'example', 'demo', 'temp', 'draft'}
        if name.lower() in generic_terms:
            return False
            
        return True
    
    def _calculate_confidence(self, sheet_name: str, client_name: str) -> float:
        """Calculate confidence score for client name detection"""
        score = 0.5  # Base score
        
        # Higher confidence for certain patterns
        if re.search(r'\(PM-[A-Z]\)', sheet_name):
            score += 0.3
        
        if re.search(r'\b(exec|executive|leadership|team|board)\b', sheet_name, re.IGNORECASE):
            score += 0.2
            
        if re.search(r'\b(meeting|agenda|contact)\b', sheet_name, re.IGNORECASE):
            score += 0.15
            
        # Company suffixes increase confidence
        if re.search(r'\b(Inc|LLC|Corp|Company|Co|Ltd|Group|Solutions|Services)\b', client_name):
            score += 0.2
            
        # Longer names tend to be more specific
        if len(client_name) > 10:
            score += 0.1
            
        return min(1.0, score)

class EnhancedExcelProcessor:
    """Enhanced Excel processor that keeps tables together while processing sheet by sheet"""
    
    def __init__(self, min_table_rows: int = 2, max_chunk_size: int = 8000):
        self.min_table_rows = min_table_rows
        self.max_chunk_size = max_chunk_size
        self.logger = logging.getLogger(__name__)
        # Initialize client detector for sheet-level client identification
        self.client_detector = SheetClientDetector()
    
    def extract_from_excel(self, doc_content: bytes, filename: str = "") -> Dict[str, Any]:
        """Extract content from Excel file with table-aware processing"""
        try:
            # Try multiple loading strategies to handle formulas properly
            try:
                # First try: Load with data_only=True and read_only=False for best formula handling
                workbook = openpyxl.load_workbook(BytesIO(doc_content), data_only=True, read_only=False)
                self.logger.info(f'Loaded Excel workbook with data_only=True, read_only=False: {len(workbook.worksheets)} sheets')
                
                # Force calculation of formulas by accessing all cells
                for sheet in workbook.worksheets:
                    for row in sheet.iter_rows():
                        for cell in row:
                            _ = cell.value  # Access value to trigger calculation
                
            except Exception as e:
                self.logger.warning(f'Failed to load with data_only=True read_only=False: {str(e)}')
                try:
                    # Second try: Standard data_only loading
                    workbook = openpyxl.load_workbook(BytesIO(doc_content), data_only=True, read_only=True)
                    self.logger.info(f'Loaded Excel workbook with data_only=True, read_only=True: {len(workbook.worksheets)} sheets')
                except Exception as e2:
                    self.logger.warning(f'Failed to load with data_only=True: {str(e2)}')
                    # Third try: Load without data_only and handle formulas manually
                    workbook = openpyxl.load_workbook(BytesIO(doc_content), read_only=True)
                    self.logger.info(f'Loaded Excel workbook without data_only: {len(workbook.worksheets)} sheets')
            
            sheets_data = {}
            
            for sheet_name in workbook.sheetnames:
                self.logger.info(f'Processing sheet: {sheet_name}')
                sheet = workbook[sheet_name]
                sheet_data = self._process_sheet_with_tables(sheet, sheet_name)
                
                self.logger.info(f'Sheet "{sheet_name}" processed: content_length={len(sheet_data.get("content", ""))}, tables={len(sheet_data.get("tables", []))}')
                
                if sheet_data["content"] and len(sheet_data["content"].strip()) > 10:  # Lowered threshold
                    sheets_data[sheet_name] = sheet_data
                    self.logger.info(f'Sheet "{sheet_name}" added to results')
                else:
                    self.logger.warning(f'Sheet "{sheet_name}" skipped - insufficient content (length: {len(sheet_data.get("content", ""))})')
            
            workbook.close()
            
            return {
                "type": "excel_sheets_enhanced",
                "sheets": sheets_data,
                "total_sheets": len(sheets_data),
                "sheet_names": list(sheets_data.keys()),
                "filename": filename
            }
            
        except Exception as e:
            self.logger.error(f'Enhanced Excel extraction failed for {filename}: {str(e)}')
            return {
                "type": "excel_error",
                "error": str(e),
                "sheets": {},
                "total_sheets": 0,
                "sheet_names": [],
                "filename": filename
            }
    
    def _process_sheet_with_tables(self, sheet, sheet_name: str) -> Dict[str, Any]:
        """Process a single sheet identifying and preserving table structures"""
        
        # Detect client information from sheet name
        sheet_client_info = self.client_detector.detect_client_from_sheet_name(sheet_name)
        
        # Get the actual used range of the sheet
        used_range = self._get_used_range(sheet)
        if not used_range:
            result = {"content": "", "tables": [], "summary": "Empty sheet"}
            if sheet_client_info:
                result["client_info"] = sheet_client_info
            return result
        
        min_row, max_row, min_col, max_col = used_range
        
        # Read all cell values into a grid
        cell_grid = self._read_cell_grid(sheet, min_row, max_row, min_col, max_col)
        
        # Identify table regions
        table_regions = self._identify_table_regions(cell_grid, min_row, min_col)
        
        # Generate content preserving table structures
        sheet_content = self._generate_sheet_content(
            sheet_name, cell_grid, table_regions, min_row, min_col
        )
        
        result = {
            "content": sheet_content,
            "tables": [self._table_region_to_dict(table) for table in table_regions],
            "summary": f"Sheet contains {len(table_regions)} table regions"
        }
        
        # Add client information if detected
        if sheet_client_info:
            result["client_info"] = sheet_client_info
            self.logger.info(f"Detected client '{sheet_client_info['sheet_client_name']}' in sheet '{sheet_name}' (confidence: {sheet_client_info['confidence']:.2f})")
        
        return result
    
    def _get_used_range(self, sheet) -> Optional[Tuple[int, int, int, int]]:
        """Get the actual used range of the sheet (ignoring empty rows/columns)"""
        try:
            min_row = sheet.min_row
            max_row = sheet.max_row
            min_col = sheet.min_column
            max_col = sheet.max_column
            
            # Check if sheet has any content
            if min_row is None or max_row is None:
                return None
                
            return (min_row, max_row, min_col, max_col)
        except:
            return None
    
    def _read_cell_grid(self, sheet, min_row: int, max_row: int, min_col: int, max_col: int) -> List[List[str]]:
        """Read all cell values into a 2D grid"""
        cell_grid = []
        
        for row_idx in range(min_row, max_row + 1):
            row_data = []
            for col_idx in range(min_col, max_col + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell_value = self._format_cell_value(cell.value)
                row_data.append(cell_value)
            cell_grid.append(row_data)
        
        return cell_grid
    
    def _format_cell_value(self, cell_value) -> str:
        """Format a cell value to string, handling formulas and errors"""
        if cell_value is None:
            return ""
        
        # Handle formula errors and references
        if isinstance(cell_value, str):
            cell_str = cell_value.strip()
            # Skip cells with formula errors
            if cell_str in ['#REF!', '#VALUE!', '#DIV/0!', '#NAME?', '#N/A', '#NULL!', '#NUM!']:
                return ""
            # Skip empty strings  
            if not cell_str:
                return ""
            return cell_str
        
        if isinstance(cell_value, (int, float)):
            if isinstance(cell_value, int):
                return str(cell_value)
            else:
                # Format floats to avoid excessive decimals
                formatted = f"{cell_value:.10f}".rstrip('0').rstrip('.')
                return formatted if formatted != "" else "0"
        elif isinstance(cell_value, datetime.datetime):
            return cell_value.strftime("%Y-%m-%d %H:%M:%S")
        elif isinstance(cell_value, datetime.date):
            return cell_value.strftime("%Y-%m-%d")
        elif isinstance(cell_value, datetime.time):
            return cell_value.strftime("%H:%M:%S")
        elif isinstance(cell_value, bool):
            return str(cell_value)
        else:
            return str(cell_value).strip()
    
    def _identify_table_regions(self, cell_grid: List[List[str]], start_row: int, start_col: int) -> List[TableRegion]:
        """Identify table regions in the cell grid"""
        if not cell_grid:
            return []
        
        table_regions = []
        visited = set()
        
        for row_idx, row in enumerate(cell_grid):
            for col_idx, cell in enumerate(row):
                if (row_idx, col_idx) in visited or not cell.strip():
                    continue
                
                # Try to identify a table starting from this cell
                table_region = self._find_table_region(
                    cell_grid, row_idx, col_idx, start_row, start_col, visited
                )
                
                if table_region:
                    table_regions.append(table_region)
                    
                    # Mark all cells in this table as visited
                    for r in range(table_region.start_row, table_region.end_row + 1):
                        for c in range(table_region.start_col, table_region.end_col + 1):
                            visited.add((r - start_row, c - start_col))
        
        return table_regions
    
    def _find_table_region(self, cell_grid: List[List[str]], start_row: int, start_col: int, 
                          sheet_start_row: int, sheet_start_col: int, visited: set) -> Optional[TableRegion]:
        """Find a table region starting from the given position"""
        
        # Check if this looks like a table header
        current_row = cell_grid[start_row]
        if not self._is_potential_header_row(current_row, start_col):
            return None
        
        # Find the extent of this table
        end_row = self._find_table_end_row(cell_grid, start_row)
        end_col = self._find_table_end_col(cell_grid, start_row, start_col)
        
        # Must have at least min_table_rows
        if end_row - start_row + 1 < self.min_table_rows:
            return None
        
        # Extract headers and data
        headers = []
        for col in range(start_col, end_col + 1):
            if col < len(current_row):
                headers.append(current_row[col].strip())
        
        # Extract data rows
        data_rows = []
        for row_idx in range(start_row + 1, end_row + 1):
            if row_idx < len(cell_grid):
                row_data = []
                for col_idx in range(start_col, end_col + 1):
                    if col_idx < len(cell_grid[row_idx]):
                        row_data.append(cell_grid[row_idx][col_idx].strip())
                data_rows.append(row_data)
        
        # Determine table type
        table_type = self._classify_table_type(headers, data_rows)
        
        return TableRegion(
            start_row=start_row + sheet_start_row,
            end_row=end_row + sheet_start_row,
            start_col=start_col + sheet_start_col,
            end_col=end_col + sheet_start_col,
            headers=headers,
            data_rows=data_rows,
            table_type=table_type
        )
    
    def _is_potential_header_row(self, row: List[str], start_col: int) -> bool:
        """Check if a row looks like a table header"""
        # Check if row has multiple non-empty cells
        non_empty_count = sum(1 for i, cell in enumerate(row[start_col:]) if cell.strip())
        return non_empty_count >= 2
    
    def _find_table_end_row(self, cell_grid: List[List[str]], start_row: int) -> int:
        """Find where the table ends vertically"""
        current_row = start_row
        consecutive_empty = 0
        
        while current_row + 1 < len(cell_grid):
            next_row = cell_grid[current_row + 1]
            non_empty_count = sum(1 for cell in next_row if cell.strip())
            
            if non_empty_count == 0:
                consecutive_empty += 1
                if consecutive_empty >= 2:  # Two empty rows = end of table
                    break
            else:
                consecutive_empty = 0
            
            current_row += 1
        
        return current_row
    
    def _find_table_end_col(self, cell_grid: List[List[str]], start_row: int, start_col: int) -> int:
        """Find where the table ends horizontally"""
        if start_row >= len(cell_grid):
            return start_col
        
        row = cell_grid[start_row]
        end_col = start_col
        
        for col_idx in range(start_col, len(row)):
            if row[col_idx].strip():
                end_col = col_idx
        
        return end_col
    
    def _classify_table_type(self, headers: List[str], data_rows: List[List[str]]) -> str:
        """Classify the type of table based on content"""
        if not headers or not data_rows:
            return "single_values"
        
        # Check for financial/numeric data
        numeric_cols = 0
        for col_idx in range(len(headers)):
            numeric_count = 0
            for row in data_rows[:5]:  # Check first 5 rows
                if col_idx < len(row):
                    cell = row[col_idx].replace(',', '').replace('$', '').replace('%', '')
                    try:
                        float(cell)
                        numeric_count += 1
                    except:
                        pass
            if numeric_count >= len(data_rows[:5]) * 0.7:  # 70% numeric
                numeric_cols += 1
        
        if numeric_cols >= len(headers) * 0.5:  # 50% of columns are numeric
            if len(data_rows) <= 10:
                return "summary_table"
            else:
                return "data_table"
        elif len(data_rows) <= 3:
            return "single_values"
        else:
            return "list"
    
    def _generate_sheet_content(self, sheet_name: str, cell_grid: List[List[str]], 
                               table_regions: List[TableRegion], start_row: int, start_col: int) -> str:
        """Generate readable content for the sheet preserving table structures"""
        content_parts = []
        content_parts.append(f"Excel Sheet: {sheet_name}")
        content_parts.append("=" * (len(sheet_name) + 14))
        
        if not table_regions:
            # No tables found, process as before but more structured
            content_parts.append("\nSheet Content:")
            for row_idx, row in enumerate(cell_grid):
                row_cells = [cell for cell in row if cell.strip()]
                if row_cells:
                    content_parts.append(f"Row {row_idx + start_row}: {' | '.join(row_cells)}")
        else:
            # Process each table region
            for i, table in enumerate(table_regions, 1):
                content_parts.append(f"\nTable {i} ({table.table_type.replace('_', ' ').title()}):")
                content_parts.append("-" * 50)
                
                # Add headers
                if table.headers:
                    header_line = " | ".join(table.headers)
                    content_parts.append(f"Headers: {header_line}")
                    content_parts.append("-" * len(header_line))
                
                # Add data rows
                for row_idx, data_row in enumerate(table.data_rows):
                    if any(cell.strip() for cell in data_row):  # Skip completely empty rows
                        row_content = " | ".join(data_row)
                        content_parts.append(f"{row_idx + 1}: {row_content}")
                
                # Add summary for larger tables
                if len(table.data_rows) > 20:
                    content_parts.append(f"... ({len(table.data_rows)} total rows in this table)")
        
        return "\n".join(content_parts)
    
    def _table_region_to_dict(self, table: TableRegion) -> Dict[str, Any]:
        """Convert TableRegion to dictionary for JSON serialization"""
        return {
            "start_row": table.start_row,
            "end_row": table.end_row,
            "start_col": table.start_col,
            "end_col": table.end_col,
            "headers": table.headers,
            "row_count": len(table.data_rows),
            "col_count": len(table.headers),
            "table_type": table.table_type
        }
    
    def chunk_excel_document(self, excel_data: Dict[str, Any], metadata: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Create chunks for Excel document, keeping tables together"""
        if excel_data.get("type") != "excel_sheets_enhanced":
            self.logger.warning(f"Invalid Excel data type: {excel_data.get('type')}")
            return []
        
        chunks = []
        sheet_index = 0
        
        for sheet_name, sheet_data in excel_data["sheets"].items():
            content = sheet_data["content"]
            tables = sheet_data["tables"]
            client_info = sheet_data.get("client_info")  # Extract client info if present
            
            if not content or len(content.strip()) < 50:  # Minimum content threshold
                continue
            
            # Check if content fits in a single chunk
            if len(content) <= self.max_chunk_size:
                # Single chunk for this sheet
                chunk = self._create_sheet_chunk(
                    content, sheet_name, sheet_index, tables, metadata, excel_data, client_info
                )
                chunks.append(chunk)
            else:
                # Split large sheets but keep tables together
                sheet_chunks = self._split_large_sheet(
                    content, sheet_name, sheet_index, tables, metadata, excel_data, client_info
                )
                chunks.extend(sheet_chunks)
            
            sheet_index += 1
        
        self.logger.info(f"Created {len(chunks)} table-aware chunks for Excel file")
        return chunks
    
    def _create_sheet_chunk(self, content: str, sheet_name: str, sheet_index: int, 
                           tables: List[Dict], metadata: Dict, excel_data: Dict, 
                           client_info: Optional[Dict] = None) -> Dict[str, Any]:
        """Create a single chunk for a sheet"""
        chunk = {
            "chunk": content.strip(),
            "chunk_id": f"{metadata['document_id']}_sheet_{sheet_index}",
            "parent_id": metadata['document_id'],
            "chunk_index": sheet_index,
            "chunk_type": "excel_sheet_enhanced",
            "sheet_name": sheet_name,
            "sheet_index": sheet_index,
            "table_count": len(tables),
            "table_types": [t["table_type"] for t in tables],
            "is_excel_sheet": True,
            "excel_sheet_count": excel_data["total_sheets"],
            **metadata
        }
        
        # Add client information if detected from sheet name
        if client_info:
            chunk.update({
                "sheet_client_name": client_info.get("sheet_client_name"),
                "sheet_client_source": client_info.get("sheet_client_source"),
                "sheet_client_confidence": client_info.get("confidence"),
                "sheet_pm_initial": client_info.get("sheet_pm_initial"),
                "sheet_pm_name": client_info.get("sheet_pm_name")
            })
        
        return chunk
    
    def _split_large_sheet(self, content: str, sheet_name: str, sheet_index: int,
                          tables: List[Dict], metadata: Dict, excel_data: Dict,
                          client_info: Optional[Dict] = None) -> List[Dict[str, Any]]:
        """Split large sheets into multiple chunks while preserving table boundaries"""
        
        # Split content by table sections
        lines = content.split('\n')
        chunks = []
        current_chunk = []
        current_size = 0
        part_index = 0
        
        table_start_lines = set()
        for table in tables:
            # Find approximate line where each table starts
            for i, line in enumerate(lines):
                if f"Table {tables.index(table) + 1}" in line:
                    table_start_lines.add(i)
                    break
        
        for i, line in enumerate(lines):
            line_size = len(line) + 1  # +1 for newline
            
            # Check if adding this line would exceed chunk size
            if current_size + line_size > self.max_chunk_size and current_chunk:
                # Don't split in the middle of a table
                if i not in table_start_lines:
                    # Create chunk
                    chunk_content = '\n'.join(current_chunk)
                    chunk = self._create_sheet_part_chunk(
                        chunk_content, sheet_name, sheet_index, part_index, 
                        metadata, excel_data, client_info
                    )
                    chunks.append(chunk)
                    
                    # Start new chunk
                    current_chunk = [line]
                    current_size = line_size
                    part_index += 1
                else:
                    # We're at a table start, include the line in current chunk anyway
                    current_chunk.append(line)
                    current_size += line_size
            else:
                current_chunk.append(line)
                current_size += line_size
        
        # Add final chunk
        if current_chunk:
            chunk_content = '\n'.join(current_chunk)
            chunk = self._create_sheet_part_chunk(
                chunk_content, sheet_name, sheet_index, part_index,
                metadata, excel_data, client_info
            )
            chunks.append(chunk)
        
        return chunks
    
    def _create_sheet_part_chunk(self, content: str, sheet_name: str, sheet_index: int,
                                part_index: int, metadata: Dict, excel_data: Dict,
                                client_info: Optional[Dict] = None) -> Dict[str, Any]:
        """Create a chunk for part of a large sheet"""
        chunk = {
            "chunk": content.strip(),
            "chunk_id": f"{metadata['document_id']}_sheet_{sheet_index}_part_{part_index}",
            "parent_id": metadata['document_id'],
            "chunk_index": f"{sheet_index}.{part_index}",
            "chunk_type": "excel_sheet_part",
            "sheet_name": sheet_name,
            "sheet_index": sheet_index,
            "part_index": part_index,
            "is_excel_sheet": True,
            "excel_sheet_count": excel_data["total_sheets"],
            **metadata
        }
        
        # Add client information if detected from sheet name
        if client_info:
            chunk.update({
                "sheet_client_name": client_info.get("sheet_client_name"),
                "sheet_client_source": client_info.get("sheet_client_source"),
                "sheet_client_confidence": client_info.get("confidence"),
                "sheet_pm_initial": client_info.get("sheet_pm_initial"),
                "sheet_pm_name": client_info.get("sheet_pm_name")
            })
        
        return chunk

# Test function
def test_excel_processor():
    """Test the enhanced Excel processor with a sample file"""
    processor = EnhancedExcelProcessor()
    
    # This would be used with actual Excel file bytes
    print("Enhanced Excel Processor initialized successfully")
    print("Key features:")
    print("- Processes sheets individually")  
    print("- Identifies and preserves table structures")
    print("- Keeps related data together in chunks")
    print("- Handles large sheets by splitting at table boundaries")
    print("- Classifies table types (data, summary, list, etc.)")

if __name__ == "__main__":
    test_excel_processor()